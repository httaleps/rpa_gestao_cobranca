import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time

# ── 1. LER O EXCEL ──────────────────────────────────────────────────
wb = openpyxl.load_workbook('../dados/clientes.xlsx')
ws = wb.active

clientes = []
for row in ws.iter_rows(min_row=2, values_only=True):  # Pula o cabeçalho
    Nome, Email, Telefone, Endereco, *_ = row
    clientes.append({
        'nome': Nome,
        'email': Email,
        'telefone': Telefone,
        'endereco': Endereco
    })

print(f"✅ {len(clientes)} clientes carregados do Excel.")

# ── 2. ABRIR NAVEGADOR E PREENCHER FORMULÁRIO ────────────────────────────────
driver = webdriver.Chrome()  # Precisa ter o ChromeDriver instalado
driver.get('http://localhost:5000')

resultados = []  # Para atualizar o Excel depois

for cliente in clientes:
    try:
        wait = WebDriverWait(driver, 10)

        # Preenche cada campo
        wait.until(EC.presence_of_element_located((By.ID, 'nome'))).clear()
        driver.find_element(By.ID, 'nome').send_keys(cliente['nome'])

        driver.find_element(By.ID, 'email').clear()  # ← adicione
        driver.find_element(By.ID, 'email').send_keys(cliente['email'])

        driver.find_element(By.ID, 'telefone').clear()  # ← adicione
        driver.find_element(By.ID, 'telefone').send_keys(str(int(cliente['telefone'])))  # ← int() evita ".0"

        driver.find_element(By.ID, 'endereco').clear()  # ← adicione
        driver.find_element(By.ID, 'endereco').send_keys(cliente['endereco'] or '')

        # Clica em cadastrar
        driver.find_element(By.CSS_SELECTOR, '#formCliente button').click()

        # Aguarda o alert e fecha
        time.sleep(1)
        alert = driver.switch_to.alert
        print(f"✅ {cliente['nome']}: {alert.text}")
        alert.accept()

        resultados.append('Cadastrado')
        time.sleep(1)

    except Exception as e:
        print(f"❌ Erro com {cliente['nome']}: {e}")
        resultados.append('Erro')

driver.quit()

# ── 3. ATUALIZAR O EXCEL COM O RESULTADO (células verdes = sucesso) ──────────
from openpyxl.styles import PatternFill

verde = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
vermelho = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

ws.cell(row=1, column=5, value='Status')  # Cabeçalho da coluna resultado

for i, status in enumerate(resultados, start=2):
    cell = ws.cell(row=i, column=5, value=status)
    cell.fill = verde if status == 'Cadastrado' else vermelho

wb.save('../dados/clientes.xlsx')
print("✅ Excel atualizado com os resultados!")