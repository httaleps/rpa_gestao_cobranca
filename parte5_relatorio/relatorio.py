import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import sqlite3
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from datetime import datetime

# ── 1. LER DO BANCO ──────────────────────────────────────────
conn = sqlite3.connect('../parte1_cadastro/instance/techsolutions.db')
df = pd.read_sql_query('''
    SELECT
        c.nome        AS Nome,
        c.telefone    AS Telefone,
        c.email       AS Email,
        f.valor       AS Valor,
        f.data_vencimento AS Vencimento,
        f.status      AS Status
    FROM fatura f
    JOIN cliente c ON f.cliente_id = c.id
''', conn)
conn.close()

print(f"✅ {len(df)} registros carregados do banco.")

dados = df.values.tolist()

# ── 2. CRIAR O EXCEL FORMATADO ───────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Relatório de Faturas"

cabecalho_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
cabecalho_font = Font(color="FFFFFF", bold=True, size=12)

borda = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

colunas = ['Nome', 'Telefone', 'E-mail', 'Valor (R$)', 'Vencimento', 'Status']

for col, titulo in enumerate(colunas, start=1):
    cell = ws.cell(row=1, column=col, value=titulo)
    cell.fill = cabecalho_fill
    cell.font = cabecalho_font
    cell.border = borda
    cell.alignment = Alignment(horizontal='center')

# ── 3. PREENCHER DADOS ───────────────────────────────────────
verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
amarelo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

for row_idx, linha in enumerate(dados, start=2):
    for col_idx, valor in enumerate(linha[:6], start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=valor)
        cell.border = borda
        cell.alignment = Alignment(horizontal='center')

    status = linha[5]

    if status == 'Enviado':
        cor = verde
    elif status == 'Falhou':
        cor = vermelho
    else:
        cor = amarelo

    for col in range(1, 7):
        ws.cell(row=row_idx, column=col).fill = cor

# ajustar largura
for col in ws.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    ws.column_dimensions[col[0].column_letter].width = max_len + 4

# ── 4. SALVAR ───────────────────────────────────────────────
arquivo = '../dados/faturas_relatorio.xlsx'
wb.save(arquivo)

print("✅ Relatório gerado com sucesso!")